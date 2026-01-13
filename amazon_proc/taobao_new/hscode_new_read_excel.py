import pandas as pd
from openpyxl import load_workbook
import datetime
import os
import DB_CON_NEW

db_con = DB_CON_NEW.Database('cn')

if __name__ == '__main__':

    print(' [--- main start ---] ' + str(datetime.datetime.now()))
    in_file = input('>> Input File : ')
    print('File name :    ', os.path.basename(in_file))
    print('Directory Name:     ', os.path.dirname(in_file))

    filepath = os.path.dirname(in_file) + str('\\') + os.path.basename(in_file)
    print('>> filepath : {}'.format(filepath))
    if os.path.isfile(filepath):
        print(">> 파일명 : {}".format(filepath))
    else:
        print(">> 파일 check")
        os._exit(1)

    # Daraframe형식으로 엑셀 파일 읽기
    wb = load_workbook(filepath)
    # ws = wb.active # 워크북 생성 시 기본적으로 생기는 워크시트를 바인딩
    a_class_df = wb['Sheet1']

    for row in range(1, a_class_df.max_row):
        cate_name_big = a_class_df.cell(row=row+1, column=1).value
        cate_name_eng = a_class_df.cell(row=row+1, column=2).value
        cate_name = a_class_df.cell(row=row+1, column=3).value
        cate_hscode = a_class_df.cell(row=row+1, column=4).value
        cate_hidden_hscode = a_class_df.cell(row=row+1, column=5).value
        print(">>({}) {} | {} | {} | {} | {}".format(row, cate_name_big, cate_name_eng, cate_name, cate_hscode, cate_hidden_hscode))

        sqlg = " select cate_hscode from goport_hscode where kbn = '1' and cate_hscode = '{}'".format(cate_hscode)
        row = db_con.selectone(sqlg)
        if not row:
            sqli = "insert into goport_hscode (cate_big, cate_name_eng, cate_name, cate_hscode, hidden_hscode, kbn) values ('{}','{}','{}','{}','{}','1')".format(cate_name_big, cate_name_eng, cate_name, cate_hscode, cate_hidden_hscode)
            db_con.execute(sqli)


    # 파일 저장 1
    # write_file_name = "c:\\project\\test_save.xlsx"
    # with pd.ExcelWriter(write_file_name) as w:
    #     a_class_df.to_excel(w, sheet_name='A시트', index=False)

    # 파일 저장 2
    # wb.save(write_file_name)

    print()
    db_con.close()
    print(' [--- main end ---] ' + str(datetime.datetime.now()))