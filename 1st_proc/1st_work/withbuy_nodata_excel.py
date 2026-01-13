import pandas as pd
from openpyxl import load_workbook
import datetime
import webbrowser
import os
import DBmodule_FR

def repStr(str):
    if str == None or str == "None":
        str = ""
    return str

if __name__ == '__main__':

    print(' [--- main start ---] ' + str(datetime.datetime.now()))

    in_file = input('>> Input File : ')
    filepath = os.path.dirname(in_file) + str('\\') + os.path.basename(in_file)
    filepath = filepath.replace("'","").replace('"','').strip()
    print('>> filepath : {}'.format(filepath))
    if filepath == "":
        print(">> {} 파일이 없습니다. (종료) ".format(in_file))
        os._exit(1)

    wb = load_workbook(filepath)
    tbl = wb['Sheet1']

    db_fs = DBmodule_FR.Database('freeship')
    sqld = "delete from withbuy_nodata "
    print(">> delete withbuy_nodata ")
    db_fs.execute(sqld)
    for row in range(1, tbl.max_row):
        center = repStr(tbl.cell(row=row+1, column=1).value)
        trackno = repStr(tbl.cell(row=row+1, column=2).value)
        orderno = repStr(tbl.cell(row=row+1, column=3).value)
        memo = repStr(tbl.cell(row=row+1, column=4).value)
        regdate = repStr(tbl.cell(row=row+1, column=5).value)
        print(">> [{}] {} | {} | {} | {} ".format(center, trackno, orderno, memo, regdate))

        sql = "select idx from withbuy_nodata where trackno = '{}'".format(trackno)
        row = db_fs.execute(sql)
        if not row:
            sqli = "insert into withbuy_nodata(center, trackno, orderno, memo, regdate) values('{}', '{}', '{}', '{}', '{}')".format(center[:30], trackno[:100], orderno[:50], memo[:500], regdate)
            db_fs.execute(sqli)
            print(">> insert : {}".format(trackno))
        else:
            sqlu = "update withbuy_nodata set center = '{}', orderno = '{}, memo = '{}', regdate = '{}' where trackno = '{}'".format(center[:30], orderno[:50], memo[:500], regdate, trackno[:100])
            db_fs.execute(sqlu)
            print(">> update : {}".format(trackno))

    print(">> table withbuy_nodata OK ")
    db_fs.close()
    view_url = "http://imp.allinmarket.co.kr/admin/goods/freeship/view_withbuy_nodata_list.asp"
    print(">> 위드바이 노데이터건 리스트 : {}".format(view_url))
    chrome_path = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe %s"
    webbrowser.open_new_tab(view_url)
    print(' [--- main end ---] ' + str(datetime.datetime.now()))
